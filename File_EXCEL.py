# This Python file uses the following encoding: utf-8
#from PyQt5 import QtCore
from File import File

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class File_EXCEL(File):
    def __init__(self):
        super().__init__()
        self.workbook = Workbook()  #EXCEL文件

    #加载文档
    def load_file(self):
        self.workbook = load_workbook(self.filepath)

    #向某单元格内写入字符串
    def write_cell_with_string(ws_string,r,c,string):
        wb = load_workbook(self.filepath)
        ws = wb[ws_string]
        ws.cell(row = r,column = c).value = string
        #QMessageBox.information(None,'消息','消息',QMessageBox.Yes | QMessageBox.No)

class excel_cell_style():
    def __init__(self):
            self.font = Font(name='华文细黑',size = 12)
            self.alignment = Alignment(horizontal = "center",vertical = "center")
            self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                     right=Side(border_style=None,color='FF000000',style = 'thin'),
                                    top=Side(border_style=None,color='FF000000',style = 'thin'),
                                     bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                    diagonal=Side(border_style=None,color='FF000000'),
                                     diagonal_direction=0,
                                     outline=Side(border_style=None,color='FF000000'),
                                     vertical=Side(border_style=None,color='FF000000'),
                                     horizontal=Side(border_style=None,color='FF000000'))
            self.fill = PatternFill("solid", fgColor="FFFFFF")

    def set_as_title(self):
        self.font = Font(name='华文细黑',size = 14,bold = True)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="FFFFFF")

    def set_as_sheet_title(self):
        self.font = Font(name='华文细黑',size = 24,bold = True)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="FFFFFF")

    def set_as_text_in_center(self):
        self.font = Font(name='华文细黑',size = 12)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="FFFFFF")

    def set_as_text_in_center_and_red(self):
        self.font = Font(name='华文细黑',size = 12)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="FF0000")

    def set_as_text_in_center_and_yellow(self):
        self.font = Font(name='华文细黑',size = 12)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="FFFF00")

    def set_as_text_in_center_and_orange(self):
        self.font = Font(name='华文细黑',size = 12)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="FFA500")

    def set_as_text_in_center_and_grey(self):
        self.font = Font(name='华文细黑',size = 12)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="BEBEBE")

    def set_as_text_in_center_and_green(self):
        self.font = Font(name='华文细黑',size = 12)
        self.alignment = Alignment(horizontal = "center",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="00CD00")

    def set_as_text_in_left(self):
        self.font = Font(name='华文细黑',size = 12)
        self.alignment = Alignment(horizontal = "left",vertical = "center")
        self.border = Border(left=Side(border_style=None,color='FF000000',style = 'thin'),
                                 right=Side(border_style=None,color='FF000000',style = 'thin'),
                                top=Side(border_style=None,color='FF000000',style = 'thin'),
                                 bottom=Side(border_style=None,color='FF000000',style = 'thin'),
                                diagonal=Side(border_style=None,color='FF000000'),
                                 diagonal_direction=0,
                                 outline=Side(border_style=None,color='FF000000'),
                                 vertical=Side(border_style=None,color='FF000000'),
                                 horizontal=Side(border_style=None,color='FF000000'))
        self.fill = PatternFill("solid", fgColor="FFFFFF")
