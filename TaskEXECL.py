# This Python file uses the following encoding: utf-8
#from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMessageBox

from openpyxl import Workbook
from openpyxl import load_workbook
from Task import Task
from LDatetime import LDateTime

class TaskEXECL():
    def __init__(self):

        self.wb = Workbook()  #普通任务列表EXCEL

    def read_task_list(self,filename,task_list):
        self.wb = load_workbook(filename)
        ws = self.wb.active
        theRow = 2
        while (ws.cell(row = theRow,column = 1).value != None):
            aTask = Task()
            col = 1
            aTask.name = ws.cell(row = theRow,column = col).value
            col += 1
            aTask.date_created.datetime = ws.cell(row = theRow,column = col).value
            col += 1
            aTask.date_planed.datetime = ws.cell(row = theRow,column = col).value
            col += 1
            aTask.date_closed.datetime = ws.cell(row = theRow,column = col).value
            col += 1
            aTask.status = ws.cell(row = theRow,column = col).value
            col += 1
            aTask.note = ws.cell(row = theRow,column = col).value
            task_list.append(aTask)
            theRow += 1
