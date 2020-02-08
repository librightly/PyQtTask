# This Python file uses the following encoding: utf-8
#from PyQt5 import QtCore
#from PyQt5 import QtWidgets

from Task import Task

from openpyxl import Workbook
from openpyxl import load_workbook

from PyQt5.QtWidgets import QMessageBox

class WorkFlowItem(Task):
    def __init__(self):        
        super().__init__()

        self.index = 0  #序号
        self.phase = ""  #阶段
        self.imput = ""  #输入
        self.output = ""  #输出
        self.responsible = ""  #负责人
        self.milestone = False  #是否使Milestone点
        self.project = ""  #所属项目

    #将一条业务流写入EXCEL文件中
    def write_a_wfi_to_EXCEL(self,filepath):
        #self.show_info()
        wb = load_workbook(filepath)
        ws = wb.active
        #QMessageBox.information(None,'消息',filepath,QMessageBox.Yes | QMessageBox.No)
        ws = wb["业务流"]
        theRow = 2
        while (ws.cell(row = theRow,column = 2).value != None):
            #QMessageBox.information(None,'消息',self.index,QMessageBox.Yes | QMessageBox.No)
            #QMessageBox.information(None,'消息',str(ws.cell(row = theRow,column = 2).value),QMessageBox.Yes | QMessageBox.No)
            if ws.cell(row = theRow,column = 2).value == self.index:
                ws.cell(row = theRow,column = 6).value = self.responsible
                ws.cell(row = theRow,column = 8).value = self.date_planed
                ws.cell(row = theRow,column = 9).value = self.date_closed
                ws.cell(row = theRow,column = 10).value = self.status
                ws.cell(row = theRow,column = 11).value = self.note                
                break
            theRow += 1
        wb.save(filepath)

    #显示业务流项信息
    def show_info(self):
        str_info = ""
        str_info = str(self.project) + "\n" + str(self.index) + "\n" + str(self.responsible) + "\n" + str(self.date_planed) + "\n" + str(self.date_closed) + "\n" + str(self.status)
        QMessageBox.information(None,'消息',str_info,QMessageBox.Yes | QMessageBox.No)
